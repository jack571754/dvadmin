from datetime import datetime
from urllib.parse import quote

import django_filters
from django.db import models
from django.db.models import Q, F
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from rest_framework import serializers
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated

from dvadmin.system.models import DownloadCenter, Users, Dept
from dvadmin.utils.filters import DataLevelPermissionsFilter
from dvadmin.utils.json_response import DetailResponse, SuccessResponse, ErrorResponse
from dvadmin.utils.request_util import get_verbose_name
from dvadmin.utils.serializers import CustomModelSerializer
from dvadmin.utils.viewset import CustomModelViewSet
from dvadmin3_flow.base_model import FlowBaseModel
from dvadmin3_flow.models import FlowData, FlowRecord, FlowAuditUsers, FlowNode, FlowInfo, AUDIT_STATUS


class FlowDataSerializer(CustomModelSerializer):
    """
    流程管理 创建/更新时的列化器
    """
    flow_info_name = serializers.CharField(source='flow_info.name',read_only=True)
    def to_representation(self, instance):
        data = super().to_representation(instance)
        data["current_node"] = instance.current_node
        data["status"] = instance.status
        data["start_user_name"] = instance.start_user and instance.start_user.name
        # 查询最新的一条记录表
        flow_record_obj = FlowRecord.objects.filter(flow_data_id=data.get('id')).order_by('-create_datetime').first()
        if flow_record_obj:
            data['pre_user'] = flow_record_obj.pre_user.all().values('id','name')
            data['pre_dept'] = flow_record_obj.pre_dept.all().values('id','name')
            data['pre_role'] = flow_record_obj.pre_role.all().values('id','name')
        # 返回已审核人信息
        data['audit_users'] = FlowAuditUsers.objects.filter(flow_record__flow_data_id=data.get('id')).values('id',
                                                                                                             'audit_user',
                                                                                                             'status',
                                                                                                             'description',
                                                                                                             'create_datetime')
        return data

    class Meta:
        model = FlowData
        fields = ['id', 'no', 'name', 'models_name', 'start_user',
                  'completed_time', 'update_datetime', 'create_datetime','status','flow_info_name']


class FlowDataCrudSerializer(CustomModelSerializer):
    """
    流程管理 创建/更新时的列化器
    """

    def to_representation(self, instance):
        data = super().to_representation(instance)
        if instance.flow_info.model_type==1:
            formData = instance.pre_change_content.get('formData')
            for key,value in formData.items():
                data[key] = value
        else:
            model_name = instance.flow_info.content_type.model
            formData = instance.pre_change_content.get('formData',{})
            for key, value in formData.items():
                data[f"{model_name}_{key}"] = value
        return data

    class Meta:
        model = FlowData
        exclude = ['pre_change_content','current_node','pre_user','pre_role','pre_dept']


class FlowDataFilter(django_filters.FilterSet):
    flow_info_name = django_filters.CharFilter(field_name='flow_info__name', lookup_expr='icontains')
    no = django_filters.CharFilter(field_name='no', lookup_expr='icontains')
    start_user_name = django_filters.CharFilter(field_name='start_user__name', lookup_expr='icontains')
    pre_user = django_filters.CharFilter(field_name='pre_user__name', lookup_expr='exact')
    pre_role = django_filters.CharFilter(field_name='pre_role__name', lookup_expr='exact')
    pre_dept = django_filters.CharFilter(field_name='pre_dept__name', lookup_expr='exact')
    status = django_filters.MultipleChoiceFilter(field_name='status', choices=AUDIT_STATUS)
    create_datetime = django_filters.DateTimeFromToRangeFilter()
    class Meta:
        model = None  # 需要在初始化时设置
        fields = []  # 初始为空，将在初始化时动态添加

    def filter_queryset(self, queryset):
        """
        重写filter_queryset方法，使得可以从query_params中解析出动态的过滤条件。
        """
        queryset = super().filter_queryset(queryset)
        request = self.request
        for key in request.query_params:
            if 'json_' in key:
                field_path = key.split('_', 1)[1]
                value = request.query_params.get(key)
                lookup_expression = f'pre_change_content__formData__{field_path}__icontains'
                queryset = queryset.filter(**{lookup_expression: value})
        return queryset


class FlowDataCrudFilter(django_filters.FilterSet):
    flow_info_name = django_filters.CharFilter(field_name='flow_info__name', lookup_expr='icontains')
    no = django_filters.CharFilter(field_name='no', lookup_expr='icontains')
    start_user_name = django_filters.CharFilter(field_name='start_user__name', lookup_expr='icontains')
    pre_user = django_filters.CharFilter(field_name='pre_user__name', lookup_expr='exact')
    pre_role = django_filters.CharFilter(field_name='pre_role__name', lookup_expr='exact')
    pre_dept = django_filters.CharFilter(field_name='pre_dept__name', lookup_expr='exact')
    status = django_filters.MultipleChoiceFilter(field_name='status', choices=AUDIT_STATUS)
    create_datetime = django_filters.DateTimeFromToRangeFilter()
    class Meta:
        model = None  # 需要在初始化时设置
        fields = []  # 初始为空，将在初始化时动态添加

    def filter_queryset(self, queryset):
        """
        重写filter_queryset方法，使得可以从query_params中解析出动态的过滤条件。
        """
        queryset = super().filter_queryset(queryset)
        request = self.request
        for key in request.query_params:
            if '_' in key:
                field_path = key.split('_', 1)[1]
                value = request.query_params.get(key)
                lookup_expression = f'pre_change_content__formData__{field_path}__icontains'
                queryset = queryset.filter(**{lookup_expression: value})
        return queryset


class FlowDataViewSet(CustomModelViewSet):
    """
    流程管理接口:
    """
    queryset = FlowData.objects.all()
    permission_classes = [IsAuthenticated]
    serializer_class = FlowDataSerializer
    export_serializer_class = FlowDataCrudSerializer
    extra_filter_class = []
    filter_class = FlowDataFilter
    search_fields = ['name']

    @action(methods=['post'], detail=False)
    def get_button_permission(self,request):
        """
        获取按钮权限
        :return:
        """
        body = request.data
        queryset = FlowData.objects.filter(models_name=body.get('models_name'),status=0).first()
        if queryset is None:
            return DetailResponse(data=[])
        data = dict()
        data["pre_user"] = queryset.pre_user.values('id', 'name')
        data["pre_role"] = queryset.pre_role.values('id', 'name')
        data["pre_dept"] = queryset.pre_dept.values('id', 'name')
        return DetailResponse(data=data)

    @action(methods=['get'], detail=False, permission_classes=[IsAuthenticated])
    def get_all_flow_data(self, request):
        """
        获取所有审批流程数据
        :param request:
        :return:
        """
        #self.extra_filter_class = []
        queryset = self.filter_queryset(self.get_queryset())
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True, request=request)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True, request=request)
        return SuccessResponse(data=serializer.data, msg="获取成功")

    @action(methods=['POST'], detail=True, permission_classes=[IsAuthenticated])
    def submit_flow_data(self, request, pk):
        """
        流程为动态表单时,提交的数据
        """
        flow_info_obj = FlowInfo.objects.filter(id=pk).first()
        if not flow_info_obj:
            return ErrorResponse(msg="流程不存在")
        flow_node_obj = FlowNode.objects.filter(flow_info=flow_info_obj, node_type='Start').first()
        users = request.user
        data = {
            "flow_info": flow_info_obj,
            "models_name": None,
            "pre_change_content": {
                "formData": request.data,
            },
            "current_node": flow_node_obj.node_dict,
            "start_user": users,
            "name": f"{users.name}发起的{flow_info_obj.name}的流程",
            "creator": users,
            "handler": users,
            "modifier": users.id,
            "dept_belong_id": users.dept_id,
        }
        flow_data_obj = FlowData.objects.create(**data)
        # 直接创建一条记录数据，并更新
        flow_record_obj = FlowRecord.objects.create(**{
            "flow_data_id": flow_data_obj.id,
            "current_node_id": flow_node_obj.node_id,
            "type": flow_node_obj.node_type,
            "status": 1,
            "parent_node_id": flow_node_obj.parent.node_id if flow_node_obj.parent else None,
            "creator": users,
            "modifier": users.id,
            "dept_belong_id": users.dept_id,
        })
        if users:
            flow_record_obj.pre_user.set([users.id])
        FlowAuditUsers.objects.create(**{
            "flow_record": flow_record_obj,
            "audit_user": flow_data_obj.start_user,
            "status": 1,
            "description": "自动审核通过",
        })
        # 调用一次执行审批流程
        FlowBaseModel.process_engine(flow_data_obj.id)
        return DetailResponse(data=flow_data_obj.id, msg="提交成功")

    @action(methods=['get'], detail=True, permission_classes=[IsAuthenticated])
    def get_pre_change_content(self,request,pk):
        """
        获取预审核数据
        """
        instance = FlowData.objects.get(id=pk)
        if instance.flow_info.model_type == 0:
            model_class = instance.flow_info.content_type.model_class()
            field_verbose_names = {}
            update_fields = instance.pre_change_content.get("update_fields")
            update_fields = instance.pre_change_content.get("update_fields")
            update_id = update_fields.get('id')
            if update_id:
                instance_obj = model_class.all_objects.filter(id=update_id).first()
                for field in model_class._meta.get_fields():
                    if hasattr(field, 'verbose_name'):
                        field_verbose_names[field.name] = field.verbose_name
                    if field.name == 'modifier':
                        users = Users.objects.filter(id=update_fields['modifier']).first()
                        if users:
                            update_fields['modifier'] = users.name
                    if field.name == 'dept_belong_id':
                        dept = Dept.objects.filter(id=update_fields['dept_belong_id']).first()
                        if dept:
                            update_fields['dept_belong_id'] = dept.name
                    if isinstance(field, models.ForeignKey):
                        related_instance = getattr(instance_obj, field.name, None)
                        if related_instance:
                            related_field_value = getattr(related_instance, 'name', None)
                            update_fields[field.name] = related_field_value

            def get_verbose_names_from_dicts(dict1, dict2):
                """将label名换为中文"""
                verbose_names = {}
                for field, value in dict2.items():
                    if field in dict1:
                        verbose_names[dict1[field]] = value
                    else:
                        verbose_names[field] = value
                return verbose_names
            verbose_names = get_verbose_names_from_dicts(field_verbose_names, update_fields)
            instance.pre_change_content["form_data"]=verbose_names
            result = instance.pre_change_content
        else:
            result = instance.pre_change_content
        return DetailResponse(data={
            'model_type':instance.flow_info.model_type,
            'pre_change_content':result,
            'formConf':instance.flow_info.form_conf
        })

    @action(methods=['get'], detail=True, permission_classes=[IsAuthenticated])
    def get_flow_record(self,request,pk):
        """
        获取审核记录
        """
        instance = FlowData.objects.get(id=pk)
        flow_record = FlowRecord.objects.filter(flow_data_id=pk).order_by("id")
        result = list()
        for record in flow_record:
            node_data = FlowNode.objects.filter(flow_info_id=instance.flow_info_id,node_id=record.current_node_id).values()
            audit_users = FlowAuditUsers.objects.filter(flow_record_id=record.id).values('description',name=F('audit_user__name'))
            result.append({
                "startUserName":instance.start_user.name,
                "startDatetime": instance.create_datetime,
                "nodeData":node_data[0] if node_data else None,
                "auditUsers":audit_users,
                "preInfo":{
                    "pre_user": record.pre_user.all().values('id', 'name'),
                    "pre_dept": record.pre_dept.all().values('id', 'name'),
                    "pre_role": record.pre_role.all().values('id', 'name'),
                },
                "nodeStatus":record.status
            })
        return DetailResponse(data=result)

    @action(methods=['get'], detail=True, permission_classes=[IsAuthenticated])
    def get_flow_process(self, request, pk):
        """
        获取流程图
        """
        instance = FlowData.objects.get(id=pk)
        result = FlowNode.get_flow_node(instance.flow_info)
        return DetailResponse(data=result)

    @action(methods=['get'], detail=False, permission_classes=[IsAuthenticated])
    def my_pending_handle(self, request):
        """
        获取待我处理数据
        :param request:
        :return:
        """
        # 0待我处理；1我已处理；2我已提交；3抄送我的
        status = int(request.query_params.get('handle_status', 0))
        filter = Q(pre_user__id=request.user.id)
        if request.user:
            filter |= Q(pre_role__id__in=list(request.user.role.all().values_list('id', flat=True)))
        if request.user.dept:
            filter |= Q(pre_dept__id=request.user.dept.id)
        queryset = self.get_queryset()
        if status == 0:
            flow_data_ids = FlowRecord.objects.filter(filter, status=0).values_list('flow_data__id', flat=True)
            filter = Q(id__in=flow_data_ids)
            self.extra_filter_class = [DataLevelPermissionsFilter]
            queryset = self.filter_queryset(queryset.filter(filter))
        elif status == 1:
            flow_record_ids = FlowAuditUsers.objects.filter(audit_user=request.user.id, status__in=[1, 2]).values_list(
                'flow_record_id', flat=True)
            flow_data_ids = FlowRecord.objects.filter(id__in=flow_record_ids, type='Approval').values_list(
                'flow_data_id', flat=True)
            queryset = queryset.filter(id__in=flow_data_ids)
        elif status == 2:
            queryset = self.filter_queryset(self.get_queryset()).filter(start_user=self.request.user)
        elif status == 3:
            filter = Q(filter) & Q(flowrecord__type='Cc')
            queryset = queryset.filter(filter)
        page = self.paginate_queryset(queryset.distinct())
        if page is not None:
            serializer = self.get_serializer(page, many=True, request=request)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True, request=request)
        return SuccessResponse(data=serializer.data, msg="获取成功")

    @action(methods=['get'], detail=False,permission_classes=[IsAuthenticated])
    def handle_to_me(self, request):
        """
        待我处理
        """
        filter = Q(pre_user__id=request.user.id)
        if request.user:
            filter |= Q(pre_role__id__in=list(request.user.role.all().values_list('id', flat=True)))
        if request.user.dept:
            filter |= Q(pre_dept__id=request.user.dept.id)
        flow_data_ids = FlowRecord.objects.filter(filter, status=0).values_list('flow_data__id', flat=True)
        filter = Q(id__in=flow_data_ids)
        # self.extra_filter_class = [DataLevelPermissionsFilter]
        queryset = self.filter_queryset(FlowData.objects.filter(filter))
        page = self.paginate_queryset(queryset.distinct())
        if page is not None:
            serializer = self.get_serializer(page, many=True, request=request)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True, request=request)
        return SuccessResponse(data=serializer.data, msg="获取成功")

    @action(methods=['get'], detail=False)
    def handle_completed(self, request):
        """
        我已处理
        """
        flow_record_ids = FlowAuditUsers.objects.filter(audit_user=request.user.id, status__in=[1, 2]).values_list(
            'flow_record_id', flat=True)
        flow_data_ids = FlowRecord.objects.filter(id__in=flow_record_ids, type='Approval').values_list(
            'flow_data_id', flat=True)
        queryset = self.get_queryset().filter(id__in=flow_data_ids)
        queryset = self.filter_queryset(queryset)
        page = self.paginate_queryset(queryset.distinct())
        if page is not None:
            serializer = self.get_serializer(page, many=True, request=request)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True, request=request)
        return SuccessResponse(data=serializer.data, msg="获取成功")

    @action(methods=['get'], detail=False)
    def handle_my_submit(self, request):
        """
        我发起的
        """
        queryset = self.get_queryset().filter(start_user=self.request.user)
        queryset = self.filter_queryset(queryset)
        page = self.paginate_queryset(queryset.distinct())
        if page is not None:
            serializer = self.get_serializer(page, many=True, request=request)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True, request=request)
        return SuccessResponse(data=serializer.data, msg="获取成功")

    @action(methods=['get'], detail=False)
    def cc_to_me(self, request):
        """
        抄送我的
        """
        filter = Q(pre_user__id=request.user.id,flowrecord__type='Cc')
        if request.user:
            filter |= Q(pre_role__id__in=list(request.user.role.all().values_list('id', flat=True)))
        if request.user.dept:
            filter |= Q(pre_dept__id=request.user.dept.id)
        filter = Q(filter)
        queryset = self.queryset.filter(filter)
        queryset = self.filter_queryset(queryset)
        page = self.paginate_queryset(queryset.distinct())
        if page is not None:
            serializer = self.get_serializer(page, many=True, request=request)
            return self.get_paginated_response(serializer.data)
        serializer = self.get_serializer(queryset, many=True, request=request)
        return SuccessResponse(data=serializer.data, msg="获取成功")

    @action(methods=['post'], detail=True, permission_classes=[IsAuthenticated])
    def handle_pass(self, request,pk, *args, **kwargs):
        """
        处理通过
        :param request:
        :return:
        """
        reason = request.data.get("reason", "无")
        instance = FlowData.objects.filter(id=pk).first()
        if instance.status != 0:
            return ErrorResponse(msg="流程数据状态错误")
        if not instance:
            return ErrorResponse(msg="数据不存在")
        if not FlowData.approval_permission(instance=instance, user=request.user):
            return ErrorResponse(msg="当前数据您无权限审核")
        # 进行审核
        flow_record_obj = FlowRecord.objects.filter(flow_data=instance.id).first()
        flow_record_obj.status = 1
        flow_record_obj.save()
        FlowAuditUsers.objects.create(**{
            "flow_record": flow_record_obj,
            "audit_user": self.request.user,
            "status": 1,
            "description": f"审核通过,处理意见:{reason}",
        })

        return DetailResponse(data=None, msg="处理通过")

    @action(methods=['post'], detail=True, permission_classes=[IsAuthenticated])
    def handle_reject(self, request, pk,*args, **kwargs):
        """
        处理驳回
        :param request:
        :return:
        """
        reason = request.data.get("reason", "无")
        instance = FlowData.objects.filter(id=pk).first()
        if instance.status != 0:
            return ErrorResponse(msg="流程数据状态错误")
        if not instance:
            return ErrorResponse(msg="数据不存在")
        if not FlowData.approval_permission(instance=instance, user=request.user):
            return ErrorResponse(msg="当前数据您无权限审核")
        # 进行审核
        flow_record_obj = FlowRecord.objects.filter(flow_data=instance.id).first()
        flow_record_obj.status = 2
        flow_record_obj.save()
        FlowAuditUsers.objects.create(**{
            "flow_record": flow_record_obj,
            "audit_user": self.request.user,
            "status": 2,
            "description": reason,
        })
        return DetailResponse(data=None, msg="处理驳回")

    @action(methods=['post'], detail=False, permission_classes=[IsAuthenticated])
    def handle_cancel(self, request, *args, **kwargs):
        """
        审核撤销
        :param request:
        :return:
        """
        instance = FlowData.objects.get(id=request.data.get('id'))
        if instance.status != 0:
            return ErrorResponse(msg="流程数据状态错误")
        filter = Q(flowrecord__pre_user=request.user.id)
        if request.user:
            filter |= Q(flowrecord__pre_role__in=list(request.user.role.all().values_list('id', flat=True)))
        if request.user.dept:
            filter |= Q(flowrecord__pre_dept=request.user.dept.id)
        if not FlowData.objects.filter(id=instance.id).filter(filter).exists():
            return ErrorResponse(msg="当前数据您无权限审核")
        # 进行审核撤销
        flow_record_obj = FlowRecord.objects.filter(flow_data=instance.id).first()
        flow_record_obj.status = 3
        flow_record_obj.save()
        FlowAuditUsers.objects.create(**{
            "flow_record": flow_record_obj,
            "audit_user": self.request.user,
            "status": 3,
            "description": "手动审核撤销",
        })
        return DetailResponse(data=None, msg="处理驳回")


    @action(methods=['get'], detail=True, permission_classes=[IsAuthenticated])
    def get_crud_columns(self, request, pk):
        """
        发起流程中查看数据的表格列设置
        """
        instance = FlowInfo.objects.filter(id=pk).first()
        if not instance:
            return ErrorResponse(msg="流程不存在")
        model_type = instance.model_type
        if model_type==0:
            conf = instance.form_conf
            components = conf.get("components")
            result = {}
            model_name = instance.content_type.model
            for component in components:
                key = component.get('key')
                title = component.get('name')
                colunmShow = component.get('columnShow')
                if colunmShow:
                    result[f"{model_name}_{key}"] = {
                        "title": title,
                        "type": "text",
                        "search": {
                            "show": True,
                        },
                        "column": {
                            "order": 0
                        }
                    }
            return DetailResponse(data=result, msg="获取成功")
        else:
            conf = instance.form_conf
            components = conf.get("components")
            result = {}
            for component in components:
                key = component.get('key')
                title = component.get('name')
                type = "text"
                search = True
                component_type = component.get('type')
                if component_type in ['ImageUpload']:
                    type = "image-uploader"
                    search = False
                if component_type in ['DateTimePicker']:
                    type = "datetime"
                    search = False

                result[f'json_{key}'] = {
                    "title": title,
                    "type": type,
                    "search": {
                        "show": search,
                    },
                    "column": {
                        "show":False
                    }
                }
                result[key] = {
                    "title":title,
                    "type":type,
                    "search":{
                        "show":False
                    },
                    "column":{
                        "order":0,
                        "width":150
                    }
                }
            return DetailResponse(data=result, msg="获取成功")

    @action(methods=['get'], detail=True, permission_classes=[IsAuthenticated])
    def get_crud_data(self, request, pk):
        """
        发起流程中的数据详情查看
        """
        instance = FlowInfo.objects.filter(id=pk).first()
        if not instance:
            return ErrorResponse(msg="流程不存在")
        model_type = instance.model_type
        queryset = FlowData.objects.filter(flow_info=instance.id)
        self.filter_class = FlowDataCrudFilter
        queryset = self.filter_queryset(queryset)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = FlowDataCrudSerializer(page, many=True, request=request)
            return self.get_paginated_response(serializer.data)
        serializer = FlowDataCrudSerializer(queryset, many=True, request=request)
        return SuccessResponse(data=serializer.data, msg="获取成功")

    @action(methods=['get'], detail=True)
    def export_data(self, request, pk):
        """
        导出功能
        :param request:
        :param args:
        :param kwargs:
        :return:
        """
        instance = FlowInfo.objects.filter(id=pk).first()
        if not instance:
            return ErrorResponse(msg="流程不存在")
        queryset = FlowData.objects.filter(flow_info=instance.id)
        queryset = self.filter_queryset(queryset)
        conf = instance.form_conf
        components = conf.get("components")
        result = {}
        for component in components:
            key = component.get('key')
            title = component.get('name')
            result[f"{key}"] = title
        self.export_field_label = result
        assert self.export_field_label, "'%s' 请配置对应的导出模板字段。" % self.__class__.__name__
        assert self.export_serializer_class, "'%s' 请配置对应的导出序列化器。" % self.__class__.__name__
        data = self.export_serializer_class(queryset, many=True, request=request).data
        try:
            from dvadmin.system.tasks import async_export_data
            async_export_data.delay(
                data,
                str(f"导出{get_verbose_name(queryset)}-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"),
                DownloadCenter.objects.create(creator=request.user,
                                              task_name=f'{get_verbose_name(queryset)}数据导出任务',
                                              dept_belong_id=request.user.dept_id).pk,
                self.export_field_label
            )
            return SuccessResponse(msg="导出任务已创建，请前往‘下载中心’等待下载")
        except:
            pass
        # 导出excel 表
        response = HttpResponse(content_type="application/msexcel")
        response["Access-Control-Expose-Headers"] = f"Content-Disposition"
        response["content-disposition"] = f'attachment;filename={quote(str(f"导出{get_verbose_name(queryset)}.xlsx"))}'
        wb = Workbook()
        ws = wb.active
        header_data = ["序号", *self.export_field_label.values()]
        hidden_header = ["#", *self.export_field_label.keys()]
        df_len_max = [self.get_string_len(ele) for ele in header_data]
        row = get_column_letter(len(self.export_field_label) + 1)
        column = 1
        ws.append(header_data)
        for index, results in enumerate(data):
            results_list = []
            for h_index, h_item in enumerate(hidden_header):
                for key, val in results.items():
                    if key == h_item:
                        if val is None or val == "":
                            results_list.append("")
                        else:
                            results_list.append(val)
                        # 计算最大列宽度
                        result_column_width = self.get_string_len(val)
                        if h_index != 0 and result_column_width > df_len_max[h_index]:
                            df_len_max[h_index] = result_column_width
            ws.append([index + 1, *results_list])
            column += 1
        # 　更新列宽
        for index, width in enumerate(df_len_max):
            ws.column_dimensions[get_column_letter(index + 1)].width = width
        tab = Table(displayName="Table", ref=f"A1:{row}{column}")  # 名称管理器
        style = TableStyleInfo(
            name="TableStyleLight11",
            showFirstColumn=True,
            showLastColumn=True,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)
        wb.save(response)
        return response